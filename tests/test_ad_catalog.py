"""Каталог рекламных конструкций: разбор файлов, схлопывание точек, ревизии.

Три вещи, которые здесь легко сломать и трудно заметить:
точка = одна конструкция (в источнике на одну координату приходится по десять
строк); пак заменяет город целиком, поэтому пустой пак применять нельзя; откат
переключает флаги, а не пересоздаёт данные.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from conftest import payload
from domain.catalog import (
    CityBounds,
    Point,
    SourceRow,
    collapse_points,
    compare_points,
)
from infrastructure.catalog.parser import ParseContext, parse_file

# Прямоугольник Севастополя из миграции — им же пользуется разбор.
SEVASTOPOL = CityBounds(44.512165, 44.687838, 33.382234, 33.649528)

LENINA_54 = (44.605398, 33.527134)
LENINA_64 = (44.604132, 33.526755)  # 144 м от Ленина 54 — ближайшая чужая точка


@pytest.fixture
def context() -> ParseContext:
    return ParseContext(city_name="Севастополь", bounds=SEVASTOPOL)


def build_xlsx(rows: list[tuple[str, ...]]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def header_and(rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    return [("Трасса/Город", "Адрес", "Координаты"), *rows]


def source_row(latitude: float, longitude: float, address: str = "адрес") -> SourceRow:
    return SourceRow(
        address=address,
        latitude=latitude,
        longitude=longitude,
        raw={"Адрес": address},
    )


# --- схлопывание точек ------------------------------------------------------


class TestCollapse:
    def test_identical_rows_become_one_structure(self):
        """Десять строк на одну координату — один щит, но десять поверхностей."""
        rows = [source_row(*LENINA_54) for _ in range(10)]

        points = collapse_points(rows)

        assert len(points) == 1
        assert points[0].surfaces_count == 10
        assert len(points[0].source_rows) == 10

    def test_close_points_merge(self):
        # Пять метров по широте — заведомо один и тот же щит.
        rows = [source_row(*LENINA_54), source_row(LENINA_54[0] + 0.000045, LENINA_54[1])]

        assert len(collapse_points(rows)) == 1

    def test_distinct_points_stay_apart(self):
        """144 метра — уже разные конструкции, слипаться они не должны."""
        rows = [source_row(*LENINA_54), source_row(*LENINA_64)]

        assert len(collapse_points(rows)) == 2

    def test_most_common_address_wins(self):
        rows = [
            source_row(*LENINA_54, address="ул. Ленина, д. 54"),
            source_row(*LENINA_54, address="ул. Ленина, д. 54"),
            source_row(*LENINA_54, address="Ленина 54"),
        ]

        assert collapse_points(rows)[0].address == "ул. Ленина, д. 54"

    def test_coordinates_come_from_first_row(self):
        """Результат не должен зависеть от того, сколько строк доложили."""
        rows = [source_row(*LENINA_54), source_row(LENINA_54[0] + 0.00002, LENINA_54[1])]

        point = collapse_points(rows)[0]

        assert (point.latitude, point.longitude) == LENINA_54

    def test_empty_input(self):
        assert collapse_points([]) == []


class TestCompare:
    def test_counts_added_removed_and_kept(self):
        previous = [Point(*LENINA_54), Point(*LENINA_64)]
        current = [Point(*LENINA_54), Point(44.548442, 33.438264)]

        diff = compare_points(previous, current)

        assert (diff.added, diff.removed, diff.kept) == (1, 1, 1)

    def test_slightly_moved_point_is_the_same_one(self):
        """Ревизии приходят от разных людей: 10 метров — то же самое место."""
        previous = [Point(*LENINA_54)]
        current = [Point(LENINA_54[0] + 0.00009, LENINA_54[1])]

        assert compare_points(previous, current).kept == 1

    def test_empty_previous_means_everything_is_new(self):
        diff = compare_points([], [Point(*LENINA_54)])

        assert (diff.added, diff.removed) == (1, 0)


class TestCityBounds:
    def test_point_inside(self):
        assert SEVASTOPOL.contains(Point(*LENINA_54))

    def test_point_in_another_city(self):
        # Ялта — до неё от границы прямоугольника десятки километров.
        assert not SEVASTOPOL.contains(Point(44.495, 34.166))

    def test_margin_lets_through_point_just_outside(self):
        just_outside = Point(SEVASTOPOL.max_latitude + 0.005, 33.5)

        assert SEVASTOPOL.contains(just_outside)


# --- разбор файлов ----------------------------------------------------------


class TestParse:
    def test_reads_screenshot_shaped_file(self, context):
        content = build_xlsx(
            header_and(
                [
                    (
                        "Севастополь",
                        "г. Севастополь, ул. Ленина , между домами 51 и 55",
                        "44.601513, 33.524612",
                    )
                ]
            )
        )

        parsed = parse_file("sev.xlsx", content, context)

        assert parsed.rejection is None
        assert len(parsed.rows) == 1
        assert parsed.rows[0].latitude == pytest.approx(44.601513)
        assert parsed.rows[0].longitude == pytest.approx(33.524612)

    def test_split_coordinate_columns_match_single_column(self, context):
        single = build_xlsx(header_and([("Севастополь", "адрес", "44.605398, 33.527134")]))
        split = build_xlsx(
            [
                ("Город", "Адрес", "Широта", "Долгота"),
                ("Севастополь", "адрес", "44.605398", "33.527134"),
            ]
        )

        first = parse_file("one.xlsx", single, context).rows[0]
        second = parse_file("two.xlsx", split, context).rows[0]

        assert (first.latitude, first.longitude) == (second.latitude, second.longitude)

    def test_swapped_coordinates_are_fixed_by_city_bounds(self, context):
        content = build_xlsx(header_and([("Севастополь", "адрес", "33.527134, 44.605398")]))

        row = parse_file("sev.xlsx", content, context).rows[0]

        assert (row.latitude, row.longitude) == pytest.approx(LENINA_54)

    def test_header_below_the_first_row(self, context):
        content = build_xlsx(
            [
                ("Реестр конструкций на ноябрь",),
                (),
                ("Трасса/Город", "Адрес", "Координаты"),
                ("Севастополь", "адрес", "44.605398, 33.527134"),
            ]
        )

        assert len(parse_file("sev.xlsx", content, context).rows) == 1

    def test_foreign_city_rejects_whole_file(self, context):
        content = build_xlsx(
            header_and(
                [
                    ("Севастополь", "свой адрес", "44.605398, 33.527134"),
                    ("Симферополь", "чужой адрес", "44.95, 34.1"),
                ]
            )
        )

        parsed = parse_file("mixed.xlsx", content, context)

        assert parsed.rejected
        assert "Симферополь" in (parsed.rejection or "")
        assert parsed.rows == []

    def test_city_prefix_does_not_make_file_foreign(self, context):
        content = build_xlsx(header_and([("г. Севастополь", "адрес", "44.605398, 33.527134")]))

        assert not parse_file("sev.xlsx", content, context).rejected

    def test_empty_city_cell_is_accepted(self, context):
        """Пак уже объявил город при загрузке — пустая ячейка не спорит с ним."""
        content = build_xlsx(header_and([("", "адрес", "44.605398, 33.527134")]))

        assert len(parse_file("sev.xlsx", content, context).rows) == 1

    def test_point_outside_city_is_dropped_with_reason(self, context):
        content = build_xlsx(header_and([("Севастополь", "Ялта", "44.495, 34.166")]))

        parsed = parse_file("sev.xlsx", content, context)

        assert parsed.rows == []
        assert "за пределами города" in parsed.row_errors[0].reason

    def test_empty_address_is_dropped(self, context):
        content = build_xlsx(header_and([("Севастополь", "", "44.605398, 33.527134")]))

        parsed = parse_file("sev.xlsx", content, context)

        assert parsed.rows == []
        assert parsed.row_errors[0].row_number == 2

    def test_raw_row_keeps_unknown_columns(self, context):
        """Файл не храним, поэтому исходная строка должна дойти до базы целиком."""
        content = build_xlsx(
            [
                ("Трасса/Город", "Адрес", "Координаты", "Владелец"),
                ("Севастополь", "адрес", "44.605398, 33.527134", "ООО Ромашка"),
            ]
        )

        row = parse_file("sev.xlsx", content, context).rows[0]

        assert row.raw["Владелец"] == "ООО Ромашка"

    def test_extra_sheets_are_reported(self, context):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in header_and([("Севастополь", "адрес", "44.605398, 33.527134")]):
            sheet.append(row)
        workbook.create_sheet("черновик")
        buffer = io.BytesIO()
        workbook.save(buffer)

        assert parse_file("sev.xlsx", buffer.getvalue(), context).extra_sheets == 1

    def test_broken_file_is_rejected_without_crashing(self, context):
        assert parse_file("sev.xlsx", b"not an excel at all", context).rejected

    def test_unsupported_extension_is_rejected(self, context):
        assert parse_file("sev.pdf", b"whatever", context).rejected

    def test_file_without_header_is_rejected(self, context):
        content = build_xlsx([("что-то", "совсем", "другое"), ("1", "2", "3")])

        assert parse_file("sev.xlsx", content, context).rejected


class TestParseCsv:
    def test_windows_1251_with_semicolon(self, context):
        """Русский Excel: кодировка 1251, разделитель `;`, координаты в кавычках."""
        text = (
            "Трасса/Город;Адрес;Координаты\r\n"
            'Севастополь;ул. Ленина, д. 54;"44.605398, 33.527134"\r\n'
        )

        parsed = parse_file("sev.csv", text.encode("windows-1251"), context)

        assert len(parsed.rows) == 1
        assert parsed.rows[0].address == "ул. Ленина, д. 54"

    def test_utf8_with_comma_delimiter(self, context):
        text = 'Трасса/Город,Адрес,Координаты\nСевастополь,адрес,"44.605398, 33.527134"\n'

        parsed = parse_file("sev.csv", text.encode("utf-8"), context)

        assert parsed.rows[0].latitude == pytest.approx(44.605398)

    def test_utf8_bom(self, context):
        text = "Трасса/Город;Адрес;Координаты\nСевастополь;адрес;44.605398 33.527134\n"

        parsed = parse_file("sev.csv", text.encode("utf-8-sig"), context)

        assert len(parsed.rows) == 1

    def test_decimal_comma(self, context):
        """«44,605398, 33,527134» — четыре куска после разделения по запятой."""
        text = 'Трасса/Город;Адрес;Координаты\nСевастополь;адрес;"44,605398, 33,527134"\n'

        parsed = parse_file("sev.csv", text.encode("utf-8"), context)

        assert parsed.rows[0].latitude == pytest.approx(44.605398)


# --- API и ревизии ----------------------------------------------------------


@pytest.fixture
def uploader(client) -> dict:
    return payload(client.post("/api/v1/users", json={"full_name": "Гончарова А."}))


@pytest.fixture
def upload(client, uploader):
    def do(rows: list[tuple[str, ...]], *, city: str = "sevastopol", name: str = "pack.xlsx"):
        return client.post(
            f"/api/v1/cities/{city}/catalog/imports",
            files=[("files", (name, build_xlsx(header_and(rows)), "application/vnd.ms-excel"))],
            data={"uploaded_by_user_id": uploader["id"]},
        )

    return do


def structures(client, city: str = "sevastopol") -> dict:
    return payload(client.get(f"/api/v1/cities/{city}/ad-structures"))


LENINA_ROWS = [
    ("Севастополь", "ул. Ленина, д. 54", "44.605398, 33.527134"),
    ("Севастополь", "ул. Ленина, д. 54", "44.605398, 33.527134"),
    ("Севастополь", "ул. Ленина, д. 64", "44.604132, 33.526755"),
]


class TestUpload:
    def test_report_before_apply_and_catalog_untouched(self, client, upload):
        report = payload(upload(LENINA_ROWS))

        assert report["points_after"] == 2
        assert report["points_before"] == 0
        assert report["collapsed_rows"] == 1
        assert report["catalog_import"]["revision"] is None
        # Пока не применили — каталог пуст, точек не видно.
        assert structures(client)["total"] == 0

    def test_row_errors_are_listed(self, client, upload):
        report = payload(
            upload(
                [
                    ("Севастополь", "ул. Ленина, д. 54", "44.605398, 33.527134"),
                    ("Севастополь", "", "44.6, 33.5"),
                ]
            )
        )

        assert report["catalog_import"]["rows_rejected"] == 1
        assert report["row_errors"][0]["reason"] == "пустой адрес"

    def test_foreign_city_file_rejects_pack(self, client, upload):
        response = upload([("Симферополь", "ул. Кирова, 1", "44.95, 34.1")])

        assert response.status_code == 400
        assert "Симферополь" in response.json()["detail"]

    def test_pack_without_usable_rows_is_refused(self, client, upload):
        """Пустой пак стёр бы город целиком — такое не применяют по ошибке."""
        response = upload([("Севастополь", "", "44.605398, 33.527134")])

        assert response.status_code == 400

    def test_unknown_city_is_404(self, client, upload):
        assert upload(LENINA_ROWS, city="atlantis").status_code == 404


class TestRevisions:
    def test_apply_makes_points_visible(self, client, upload):
        report = payload(upload(LENINA_ROWS))

        applied = payload(
            client.post(f"/api/v1/catalog/imports/{report['catalog_import']['id']}/apply")
        )

        assert applied["revision"] == 1
        assert applied["is_current"] is True
        assert structures(client)["total"] == 2

    def test_second_pack_replaces_the_first(self, client, upload):
        first = payload(upload(LENINA_ROWS))
        client.post(f"/api/v1/catalog/imports/{first['catalog_import']['id']}/apply")

        second = payload(
            upload([("Севастополь", "Фиолентовое шоссе", "44.548442, 33.438264")])
        )
        applied = payload(
            client.post(f"/api/v1/catalog/imports/{second['catalog_import']['id']}/apply")
        )

        assert applied["revision"] == 2
        assert second["points_before"] == 2
        assert second["removed"] == 2
        listed = structures(client)
        assert listed["total"] == 1
        assert listed["items"][0]["address"] == "Фиолентовое шоссе"

    def test_restore_returns_previous_revision(self, client, upload):
        first = payload(upload(LENINA_ROWS))
        first_id = first["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{first_id}/apply")
        second = payload(
            upload([("Севастополь", "Фиолентовое шоссе", "44.548442, 33.438264")])
        )
        client.post(f"/api/v1/catalog/imports/{second['catalog_import']['id']}/apply")

        restored = payload(client.post(f"/api/v1/catalog/imports/{first_id}/restore"))

        # Откат не заводит новую ревизию: снова показывается первая.
        assert restored["revision"] == 1
        assert restored["is_current"] is True
        assert structures(client)["total"] == 2

    def test_applying_twice_conflicts(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")

        assert client.post(f"/api/v1/catalog/imports/{import_id}/apply").status_code == 409

    def test_restoring_current_revision_conflicts(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")

        assert (
            client.post(f"/api/v1/catalog/imports/{import_id}/restore").status_code == 409
        )

    def test_hiding_the_only_revision_empties_the_catalog(self, client, upload):
        """Единственная ревизия города должна сниматься с показа.

        Иначе она несъёмная: откатиться не на что, а удалять текущую нельзя, —
        та же дверь в одну сторону, что была у «удаления» городов.
        """
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")

        hidden = payload(client.post(f"/api/v1/catalog/imports/{import_id}/hide"))

        assert hidden["is_current"] is False
        # Номер ревизии остаётся: она снята с показа, а не отменена.
        assert hidden["revision"] == 1
        assert structures(client)["total"] == 0

    def test_hidden_revision_comes_back_by_restore(self, client, upload):
        """Возврат делается обычным откатом — отдельной ручки быть не должно."""
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")
        client.post(f"/api/v1/catalog/imports/{import_id}/hide")

        restored = payload(client.post(f"/api/v1/catalog/imports/{import_id}/restore"))

        assert restored["is_current"] is True
        assert restored["revision"] == 1
        assert structures(client)["total"] == 2

    def test_hidden_revision_can_be_deleted(self, client, upload):
        """Снятая с показа удаляется как обычная старая — запрет снимается вместе с флагом."""
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")
        client.post(f"/api/v1/catalog/imports/{import_id}/hide")

        assert client.delete(f"/api/v1/catalog/imports/{import_id}").status_code == 204
        assert payload(client.get("/api/v1/cities/sevastopol/catalog/imports")) == []

    def test_hiding_what_is_not_shown_conflicts(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]

        assert client.post(f"/api/v1/catalog/imports/{import_id}/hide").status_code == 409

    def test_current_revision_cannot_be_deleted(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]
        client.post(f"/api/v1/catalog/imports/{import_id}/apply")

        assert client.delete(f"/api/v1/catalog/imports/{import_id}").status_code == 409

    def test_unapplied_pack_can_be_cancelled(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        import_id = report["catalog_import"]["id"]

        assert client.delete(f"/api/v1/catalog/imports/{import_id}").status_code == 204
        assert payload(client.get("/api/v1/cities/sevastopol/catalog/imports")) == []

    def test_history_lists_uploader(self, client, upload, uploader):
        upload(LENINA_ROWS)

        history = payload(client.get("/api/v1/cities/sevastopol/catalog/imports"))

        assert history[0]["uploaded_by"]["id"] == uploader["id"]
        assert history[0]["file_names"] == ["pack.xlsx"]

    def test_other_city_is_untouched(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        client.post(f"/api/v1/catalog/imports/{report['catalog_import']['id']}/apply")

        assert structures(client, "simferopol")["total"] == 0


class TestStructuresList:
    def test_search_by_address(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        client.post(f"/api/v1/catalog/imports/{report['catalog_import']['id']}/apply")

        found = payload(
            client.get("/api/v1/cities/sevastopol/ad-structures", params={"search": "64"})
        )

        assert found["total"] == 1
        assert found["items"][0]["address"] == "ул. Ленина, д. 64"

    def test_surfaces_count_is_visible(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        client.post(f"/api/v1/catalog/imports/{report['catalog_import']['id']}/apply")

        listed = structures(client)["items"]
        by_address = {item["address"]: item for item in listed}

        assert by_address["ул. Ленина, д. 54"]["surfaces_count"] == 2
        assert by_address["ул. Ленина, д. 64"]["surfaces_count"] == 1

    def test_pagination(self, client, upload):
        report = payload(upload(LENINA_ROWS))
        client.post(f"/api/v1/catalog/imports/{report['catalog_import']['id']}/apply")

        page = payload(
            client.get(
                "/api/v1/cities/sevastopol/ad-structures",
                params={"page": 2, "page_size": 1},
            )
        )

        assert page["total"] == 2
        assert len(page["items"]) == 1
